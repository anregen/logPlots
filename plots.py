from bokeh.plotting import figure, output_file, show
from bokeh.models import HoverTool, BoxSelectTool
from bokeh.layouts import column
from openpyxl import load_workbook
import numpy


def get_array(cell_iterable):
    ra = []
    for i in cell_iterable:
        ra.append(i.value)
    return ra

wb = load_workbook('/home/carl/Downloads/Incident_ProcessedLog No Plots.xlsx')
names = wb.get_sheet_names()
ws = wb['3009']
time_ms_3009 = get_array(ws['A'])
timestamp_ms_3009 = get_array(ws['B'])
tas_kts_3009 = get_array(ws['E'])
ias_kts_3009 = get_array(ws['F'])
ground_speed_kts_3009 = get_array(ws['T'])

ws = wb['3010']
# fgh roll pitch yaw
time_ms_3010 = get_array(ws['A'])
timestamp_ms_3010 = get_array(ws['B'])
roll_rate_3010 = get_array(ws['F'])
pitch_rate_3010 = get_array(ws['G'])
yaw_rate_3010 = get_array(ws['H'])

# mnopqr, roll pitch head, rate, rate, rate
ws = wb['4000']
time_ms_4000 = get_array(ws['A'])
timestamp_ms_4000 = get_array(ws['B'])
roll_4000 = get_array(ws['M'])
pitch_4000 = get_array(ws['N'])
head_4000 = get_array(ws['O'])
roll_rate_4000 = get_array(ws['P'])
pitch_rate_4000 = get_array(ws['Q'])
yaw_rate_4000 = get_array(ws['R'])

# cdem lat lng rotcmd, towrel
ws = wb['55002']
time_ms_55002 = get_array(ws['A'])
timestamp_ms_55002 = get_array(ws['B'])
lat_cmd_55002 = get_array(ws['C'])
lon_cmd_55002 = get_array(ws['D'])
rot_cmd_55002 = get_array(ws['E'])
towrel_cmd_55002 = get_array(ws['M'])

# HIJK radalt, snr, rtkalt, altsource
ws = wb['56003']
time_ms_56003 = get_array(ws['A'])
timestamp_ms_56003 = get_array(ws['B'])
radalt_56003 = get_array(ws['H'])
radalt_snr_56003 = get_array(ws['I'])
rtkalt_56003 = get_array(ws['J'])
rtk_model_56003 = get_array(ws['K'])

# cdefghst rudc rudp elc elp ailc ailp trc trp
ws = wb['56101']
time_ms_56101 = get_array(ws['A'])
timestamp_ms_56101 = get_array(ws['B'])
rud_cmd_56101 = get_array(ws['C'])
rud_pos_56101 = get_array(ws['D'])
elv_cmd_56101 = get_array(ws['E'])
elv_pos_56101 = get_array(ws['F'])
ail_cmd_56101 = get_array(ws['G'])
ail_pos_56101 = get_array(ws['H'])
towrel_cmd_56101 = get_array(ws['S'])
towrel_pos_56101 = get_array(ws['T'])


# output to static HTML file
output_file("plots.html")
hover = HoverTool(
    tooltips=[
        ('index', '$index'),
        ('time', '@x{0.000}'),
        ('value', '@y{0.0}')
    ],
    mode='vline'
)
# create a new plot with a title and axis labels
p1 = figure(title="Speed", x_axis_label='time (ms)', y_axis_label='kts')
p1.add_tools(hover)
p1.add_tools(BoxSelectTool(dimensions="width"))
p1.line(time_ms_3009[2:], ground_speed_kts_3009[2:], legend="Ground Speed", line_color='blue')
p1.line(time_ms_3009[2:], tas_kts_3009[2:], legend="True Air Speed", line_color='red')
p1.line(time_ms_3009[2:], ias_kts_3009[2:], legend="Indicated Air Speed", line_color='orange')

p2 = figure(title="Pitch", x_range=p1.x_range)
p2.add_tools(hover)
p2.add_tools(BoxSelectTool(dimensions="width"))
p2.line(time_ms_4000[2:], pitch_4000[2:], legend="Pitch")

p3 = figure(title="Pitch Control", x_range=p1.x_range)
p3.add_tools(hover)
p3.add_tools(BoxSelectTool(dimensions="width"))
lon_cmd_flip = [-1*x for x in lon_cmd_55002[2:]]
p3.line(time_ms_55002[2:], lon_cmd_flip, legend="-Longitude Cmd (55002)", line_color='blue')
p3.line(time_ms_56101[2:], elv_cmd_56101[2:], legend="Elevator cmd (56101)", line_color='red')
p3.line(time_ms_56101[2:], elv_pos_56101[2:], legend="Elevator pos (56101)", line_color='orange')

pitch_rate_4000_d = [x*180/numpy.pi for x in pitch_rate_4000[2:]]
pitch_rate_3010_d = [x*180/numpy.pi for x in pitch_rate_3010[2:]]
p4 = figure(title="Pitch Rate", x_range=p1.x_range)
p4.add_tools(hover)
p4.add_tools(BoxSelectTool(dimensions="width"))
p4.line(time_ms_3010[2:], pitch_rate_3010_d, legend="Rate (3010)", line_color='blue')
p4.line(time_ms_4000[2:], pitch_rate_4000_d, legend = "Rate (4000)", line_color="red")

show(column(p1, p2, p3, p4, sizing_mode='stretch_both'))
